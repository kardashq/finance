from datetime import datetime, timedelta

from django.db.models import Sum
from django.http import HttpResponse
from rest_framework import generics, viewsets, mixins
from rest_framework.authentication import TokenAuthentication
from rest_framework.generics import get_object_or_404
from rest_framework.pagination import LimitOffsetPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Transaction, Account
from .serializers import AccountSerializer, ActionSerializer, TransactionSerializer


class MyPageNumberPagination(LimitOffsetPagination):
    default_limit = 5


class TransactionsAPIView(generics.ListAPIView):
    """Endpoint for view all transactions for auth user only"""
    queryset = Transaction.objects.all()
    authentication_classes = (TokenAuthentication,)
    serializer_class = TransactionSerializer
    permission_classes = (IsAuthenticated,)
    pagination_class = MyPageNumberPagination

    def get_queryset(self):
        """Return object for current authenticated user only"""
        return self.queryset.filter(account__user=self.request.user)


class AccountDetail(generics.RetrieveUpdateAPIView):
    """Endpoint for get and put, auth user is used
    for filter"""
    serializer_class = AccountSerializer
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    queryset = Account.objects.all()

    def get_object(self):
        return self.queryset.filter(user=self.request.user).first()


class ActionViewSet(viewsets.GenericViewSet,
                    mixins.RetrieveModelMixin,
                    mixins.CreateModelMixin,
                    mixins.UpdateModelMixin,
                    mixins.DestroyModelMixin):
    """Endpoint for CRUD operation for transaction"""
    serializer_class = ActionSerializer
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    queryset = Transaction.objects.all()
    pagination_class = MyPageNumberPagination

    def perform_create(self, serializer):
        """Create new transaction"""
        account = get_object_or_404(Account, user=self.request.user)
        return serializer.save(account=account)

    def get_queryset(self):
        """Return object for current authenticated user only"""
        return self.queryset.filter(account__user=self.request.user)


class ExportTransactionsView(APIView):
    """For download Excel account statistics"""
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        transactions = Transaction.objects.filter(account__user=user)

        # Create new workbook and get the active sheet
        wb = Workbook()
        ws = wb.active

        # headers settings
        headers = [
            'ID', 'Описание', 'Категория', 'Сумма', 'Тип', 'Баланс'
        ]
        for idx, header in enumerate(headers):
            column = get_column_letter(idx + 1)
            ws.column_dimensions[column].width = 20
            ws[f'{column}1'] = header

        # Filling the table with data
        for idx, transaction in enumerate(transactions):
            row = idx + 2
            ws[f'A{row}'] = transaction.id
            ws[f'B{row}'] = transaction.description
            ws[f'C{row}'] = transaction.category.title
            ws[f'D{row}'] = transaction.amount
            ws[f'E{row}'] = 'Доход' if transaction.type_of_transaction == 'income' else 'Расход'
        ws[f'F2'] = transactions[0].account.balance

        # Preparing an HTTP-response with the created Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={user.username}-transactions.xlsx'
        wb.save(response)

        return response


class MonthlyStatsView(generics.ListAPIView):
    """Get monthly statistics on expenses for auth user only.
     Need query_params month, year and type of transaction[t](optional)"""
    serializer_class = TransactionSerializer
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    pagination_class = MyPageNumberPagination

    def get_queryset(self):
        # Get the month and year from the query
        month = self.request.query_params.get('month')
        year = self.request.query_params.get('year')
        t = self.request.query_params.get('t', None)

        # start and end date of the month
        start_date = datetime(int(year), int(month), 1)
        end_date = start_date.replace(month=start_date.month + 1, day=1) - timedelta(days=1)

        if t:
            # Get monthly statistics selected type "income" or "expense"
            stats = Transaction.objects.filter(
                account__user=self.request.user,
                type_of_transaction=t,
                date__range=[start_date, end_date]
            ).annotate(total=Sum('amount'))

            return stats
        else:
            # Get monthly statistics
            stats = Transaction.objects.filter(
                account__user=self.request.user,
                date__range=[start_date, end_date]
            ).annotate(total=Sum('amount'))

            return stats
